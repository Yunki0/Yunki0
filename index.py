import tkinter as tk
import openpyxl
from tkinter import messagebox, Menu, FLAT


#Les fonctions

# fontion qui permet de valider les infos dans les champs remplis
def valider_infos():
    nom = nom_entry.get()
    edition = edition_entry.get()
    classe = classe_entry.get()
    date = date_entry.get()
    quantite = quantite_entry.get()

    if nom and edition and classe or date or quantite:
        if quantite.isdigit():
            quantite = int(quantite)
            exporter_excel(nom, edition, classe, date, quantite)
            messagebox.showinfo("Succès", "Informations exportées avec succès.")
            nom_entry.delete(0, tk.END)
            edition_entry.delete(0, tk.END)
            classe_entry.delete(0, tk.END)
            date_entry.delete(0, tk.END)
            quantite_entry.delete(0, tk.END)
        else:
            messagebox.showerror("Erreur", "La quantité doit être un entier.")
    else:
        messagebox.showerror("Erreur", "Veuillez remplir tous les champs.")

# fonction qui permet d'exporter les éléments des champs remplis dans le fichiers excel
def exporter_excel(nom, edition, classe, date, quantite):
    wb = openpyxl.load_workbook("livres.xlsx")
    feuille = wb.active

    derniere_ligne = feuille.max_row + 1
    feuille.cell(row=derniere_ligne, column=1, value=nom)
    feuille.cell(row=derniere_ligne, column=2, value=edition)
    feuille.cell(row=derniere_ligne, column=3, value=classe)
    feuille.cell(row=derniere_ligne, column=4, value=date)
    feuille.cell(row=derniere_ligne, column=5, value=quantite)

    wb.save("livres.xlsx")

# fonction qui permet de vider le fichiers excel avec confirmation
def vider_fichier():
    confirmation = messagebox.askyesno("Confirmation", "Êtes-vous sûr(e) de vouloir effacer le contenu du fichier ?")
    if confirmation:
        wb = openpyxl.Workbook()
        wb.save("livres.xlsx")



# Création de la fenetre principale
fenetre = tk.Tk()
fenetre.title("CherryBlossoms")
fenetre.config(background='#012970')


# Création de la barre de menu
menu_bar = Menu(fenetre, bg='#fefefe')
fenetre.config(menu=menu_bar)

# Création du menu
file_menu = Menu(menu_bar, tearoff=0)
menu_bar.add_cascade(label="Fichier", menu=file_menu)
file_menu.add_command(label="...")

# Creation des champs
nom_label = tk.Label(fenetre, text="Nom du livre:", fg='white', bg='#012970')
edition_label = tk.Label(fenetre, text="Edition:", fg='white', bg='#012970')
classe_label = tk.Label(fenetre, text="Classe:", fg='white', bg='#012970')
date_label = tk.Label(fenetre, text="Date de parution:", fg='white', bg='#012970')
quantite_label = tk.Label(fenetre, text="Quantité:", fg='white', bg='#012970')

nom_entry = tk.Entry(fenetre)
edition_entry = tk.Entry(fenetre)
classe_entry = tk.Entry(fenetre)
date_entry = tk.Entry(fenetre)
quantite_entry = tk.Entry(fenetre)

valider_button = tk.Button(fenetre, text="Valider", relief=FLAT, command=valider_infos, fg='white', bg='#012970')
vider_button = tk.Button(fenetre, text="Vider", relief=FLAT, command=vider_fichier, fg='white', bg='#012970')

# Placement des widgets à l'aide de la gestion de mise en page grid
nom_label.grid(row=0, column=0, padx=10, pady=10)
nom_entry.grid(row=0, column=1, padx=10, pady=10)

edition_label.grid(row=1, column=0, padx=10, pady=10)
edition_entry.grid(row=1, column=1, padx=10, pady=10)

classe_label.grid(row=2, column=0, padx=10, pady=10)
classe_entry.grid(row=2, column=1, padx=10, pady=10)

date_label.grid(row=3, column=0, padx=10, pady=10)
date_entry.grid(row=3, column=1, padx=10, pady=10)

quantite_label.grid(row=4, column=0, padx=10, pady=10)
quantite_entry.grid(row=4, column=1, padx=10, pady=10)

valider_button.grid(row=5, columnspan=2, padx=10, pady=10)
vider_button.grid(row=5, column=1,  padx=10, pady=10)


# boucle
fenetre.mainloop()